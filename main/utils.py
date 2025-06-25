
from typing import List
import pandas as pd
import re
import imaplib
import email
from email.header import decode_header
from datetime import datetime,timedelta
import os
from django.conf import settings
import pdfplumber
from .models  import InvoiceInfo,ColumnMapping,CaseModel,WurthReport,McGrathReport,YhiaustraliaReport,RepcoReport,Supplier
from .service_repository import PurchaseReportServices
import random
from django.core.cache import cache
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.utils import parsedate_to_datetime
import shutil
from openpyxl import load_workbook

class MailAutomationClass:
    _USERNAME = 'praveenintellect6@gmail.com'
    _PASSWORD = 'ldgw urxs hylr vqib'
    # _USERNAME = 'intellectorderprocessing@gmail.com'
    # _PASSWORD = 'dkxp yqfr idgn vkqy'
    _BASE_DIRECTORY= settings.INPUT_INVOICE
    _OUT_BASE_DIRECTORY= settings.OUTPUT_INVOICE

    @staticmethod
    def send_mail(to_email, subject, body):
        sender_email = MailAutomationClass._USERNAME
        app_password = MailAutomationClass._PASSWORD
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))
        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                server.login(sender_email, app_password)
                server.send_message(msg)
            print("Email sent successfully!")
        except Exception as e:
            print("Error:", e)                

    @staticmethod
    def fetch_emails_on_date(target_date):
        def generate_random_3_digit():
            return random.randint(100, 999)
        emails = []
        attachments=[]
        folder_date_path=''
        subject_name=dict()
        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(MailAutomationClass._USERNAME, MailAutomationClass._PASSWORD)
            mail.select("inbox")
            since_date = target_date.strftime("%d-%b-%Y")
            #since_date = target_date
            before_date = (target_date + timedelta(days=1)).strftime("%d-%b-%Y")
            status, messages = mail.search(None, 'SINCE', since_date, 'BEFORE', before_date)
            if status != "OK":
                print("No emails found.")
                mail.logout()
                return []
            email_ids = messages[0].split()
            for eid in email_ids:
                try:
                    _, msg_data = mail.fetch(eid, '(RFC822)')
                    if not msg_data or not msg_data[0]:
                        continue
                    raw_email = msg_data[0][1]
                    msg = email.message_from_bytes(raw_email)
                    body = ""
                    if msg.is_multipart():
                        if "invoice" in msg.get("Subject", "").lower():
                            for part in msg.walk():
                                content_type = part.get_content_type()
                                content_disposition = str(part.get("Content-Disposition", "")).lower()
                                if part.get_filename():
                                    filename = part.get_filename()
                                    filename=str(generate_random_3_digit())+filename
                                    # Generate filename if missing
                                    if not filename:
                                        ext = mimetypes.guess_extension(content_type) or ".bin"
                                        filename = f"attachment_{uuid.uuid4().hex}{ext}"    
                                    os.makedirs(MailAutomationClass._BASE_DIRECTORY, exist_ok=True)
                                    os.makedirs(MailAutomationClass._OUT_BASE_DIRECTORY,exist_ok=True)
                                    date_str = target_date.strftime("%Y-%m-%d")
                                    folder_path = os.path.join(MailAutomationClass._BASE_DIRECTORY, date_str)
                                    folder_date_path=folder_path
                                    os.makedirs(folder_path, exist_ok=True)
                                    filepath = os.path.join(folder_path, filename)
                                    try:
                                        with open(filepath, "wb") as f:
                                            f.write(part.get_payload(decode=True))
                                        attachments.append(filepath)
                                        # print("Saved attachment:", filepath)
                                        subject_name[filename]=msg.get("Subject", "")
                                    except Exception as save_error:
                                        print(f"Failed to save attachment {filename}: {save_error}")

                                # elif content_type == "text/plain" and not body:
                                #     body = part.get_payload(decode=True).decode(errors="ignore")
                                #         else:
                                #             body = msg.get_payload(decode=True).decode(errors="ignore")

                    emails.append({
                        "from": msg.get("From"),
                        "subject": msg.get("Subject"),
                        "date": msg.get("Date"),
                        "body": body
                    })
                except Exception as e:
                    print(f"Error reading message {eid.decode(errors='ignore')}: {e}")
            mail.logout()
        except Exception as e:
            print(f"Error connecting to Gmail: {e}")
        return True,subject_name,folder_date_path,attachments
    
   
    
    @staticmethod
    def fetch_emails_cron_task(target_date):
        count=0
        basepath=os.path.join('media',"in_invoice",target_date.strftime("%Y-%m-%d"))
        os.makedirs(basepath,exist_ok=True)
        file_count = len([f for f in os.listdir(basepath) if os.path.isfile(os.path.join(basepath, f))])
        def generate_random_3_digit():
            return random.randint(100, 999)
        emails = []
        attachments=[]
        folder_date_path=''
        subject_name=dict()
        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(MailAutomationClass._USERNAME, MailAutomationClass._PASSWORD)
            mail.select("inbox")
            since_date = target_date.strftime("%d-%b-%Y")
            #since_date = target_date
            before_date = (target_date + timedelta(days=1)).strftime("%d-%b-%Y")
            # status, messages = mail.search(None, 'SINCE', since_date, 'BEFORE', before_date)
            status, messages = mail.search(None,'UNSEEN')
            if status != "OK":
                print("No emails found.")
                mail.logout()
                return []
            email_ids = messages[0].split()
            for eid in email_ids:
                try:
                    _, msg_data = mail.fetch(eid, '(RFC822)')
                    if not msg_data or not msg_data[0]:
                        continue
                    raw_email = msg_data[0][1]
                    msg = email.message_from_bytes(raw_email)
                    body = ""
                    if msg.is_multipart():
                        if "invoice" in msg.get("Subject", "").lower():
                            for part in msg.walk():
                                content_type = part.get_content_type()
                                content_disposition = str(part.get("Content-Disposition", "")).lower()
                                if part.get_filename():
                                    count+=1
                                    if count>file_count:
                                        filename = part.get_filename()
                                        filename=str(generate_random_3_digit())+filename
                                        # Generate filename if missing
                                        if not filename:
                                            ext = mimetypes.guess_extension(content_type) or ".bin"
                                            filename = f"attachment_{uuid.uuid4().hex}{ext}"    
                                        os.makedirs(MailAutomationClass._BASE_DIRECTORY, exist_ok=True)
                                        os.makedirs(MailAutomationClass._OUT_BASE_DIRECTORY,exist_ok=True)
                                        date_str = target_date.strftime("%Y-%m-%d")
                                        folder_path = os.path.join(MailAutomationClass._BASE_DIRECTORY, date_str)
                                        folder_date_path=folder_path
                                        os.makedirs(folder_path, exist_ok=True)

                                        filepath = os.path.join(folder_path, filename)
                                        try:
                                            with open(filepath, "wb") as f:
                                                f.write(part.get_payload(decode=True))
                                            attachments.append(filepath)
                                            # print("Saved attachment:", filepath)
                                            subject_name[filename]=msg.get("Subject", "")
                                        except Exception as save_error:
                                            print(f"Failed to save attachment {filename}: {save_error}")

                                # elif content_type == "text/plain" and not body:
                                #     body = part.get_payload(decode=True).decode(errors="ignore")
                                #         else:
                                #             body = msg.get_payload(decode=True).decode(errors="ignore")

                    emails.append({
                        "from": msg.get("From"),
                        "subject": msg.get("Subject"),
                        "date": msg.get("Date"),
                        "body": body
                    })
                except Exception as e:
                    print(f"Error reading message {eid.decode(errors='ignore')}: {e}")
            mail.logout()
        except Exception as e:
            print(f"Error connecting to Gmail: {e}")
        return True,subject_name,folder_date_path,attachments,count
    

    # def mail_unseen_task():
    #     try:
    #         mail = imaplib.IMAP4_SSL("imap.gmail.com")
    #         mail.login(MailAutomationClass._USERNAME, MailAutomationClass._PASSWORD)
            
    #         # Select the 'All Mail' folder only once
    #         mail.select('"[Gmail]/All Mail"')
            
    #         # Search for unread messages using Gmail's X-GM-RAW query
    #         status, message_numbers = mail.search(None, 'X-GM-RAW', 'is:unread')
            
    #         if status != "OK":
    #             print("No unread emails found.")
    #             mail.logout()
    #             return
            
    #         for num in message_numbers[0].split():
    #             # Fetch only the FLAGS to check if message is unread (no \Seen flag)
    #             status, msg_data = mail.fetch(num, '(FLAGS)')
    #             if status != "OK":
    #                 continue
                
    #             flags = msg_data[0].decode()
    #             # If \Seen flag is not present, email is unread
    #             if '\\Seen' not in flags:
    #                 # Fetch minimal headers without marking read
    #                 status, msg_data = mail.fetch(num, '(BODY.PEEK[HEADER])')
    #                 if status != "OK":
    #                     continue
                    
    #                 msg = email.message_from_bytes(msg_data[0][1])
    #                 subject, encoding = decode_header(msg.get("Subject", ""))[0]
    #                 if isinstance(subject, bytes):
    #                     subject = subject.decode(encoding or 'utf-8')
    #                 print(f"Unread email subject: {subject}")
    #                 # Process your email or attachments here without marking it read
    #         mail.logout()
    #     except Exception as e:
    #         print(f"Error: {e}")

    @staticmethod
    def mail_unseen_task():
        def generate_random_3_digit():
            return random.randint(100, 999)
        emails = []
        attachments=[]
        folder_date_path=''
        subject_name=dict()
        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(MailAutomationClass._USERNAME, MailAutomationClass._PASSWORD)
            mail.select("inbox")
            mail.select('"[Gmail]/All Mail"')
            status, messages = mail.search(None, 'X-GM-RAW', 'is:unread')
            print("status",status)
            if status != "OK":
                print("No emails found.")
                mail.logout()
            else:
                for num in messages[0].split():
                    res, data = mail.fetch(num, "(RFC822)")
                    msg = email.message_from_bytes(data[0][1])
                    subject = decode_header(msg["Subject"])[0][0]
                    try:
                        _, msg_data = mail.fetch(num, '(RFC822)')
                        if not msg_data or not msg_data[0]:
                            continue
                        raw_email = msg_data[0][1]
                        msg = email.message_from_bytes(raw_email)
                        body = ""
                        if msg.is_multipart():
                            if "invoice" in msg.get("Subject", "").lower():
                                for part in msg.walk():
                                    content_type = part.get_content_type()
                                    content_disposition = str(part.get("Content-Disposition", "")).lower()
                                    if content_disposition and "attachment" in content_disposition:
                                    # if  part.get_filename():
                                        raw_date = msg.get("Date")
                                        parsed_date = parsedate_to_datetime(raw_date).date()
                                        date_str =  parsed_date.strftime("%y-%m-%d")
                                        date_obj = datetime.strptime(date_str, "%y-%m-%d")
                                        year = str(date_obj.year)
                                        basepath=os.path.join('media',year)
                                        os.makedirs(basepath,exist_ok=True)
                                        # remote_path=r"\\system8\D\Oakley\praveen"
                                        #remote_basepath=os.path.join(remote_path,year)
                                        #os.makedirs(remote_basepath,exist_ok=True)
                                        month = date_obj.strftime("%b")
                                        basepath=os.path.join(basepath,month)
                                        os.makedirs(basepath,exist_ok=True)
                                        #remote_basepath=os.path.join(remote_basepath,month)
                                        #os.makedirs(remote_basepath,exist_ok=True)
                                        day = date_str
                                        basepath=os.path.join(basepath,day)
                                        os.makedirs(basepath,exist_ok=True)
                                        #remote_basepath=os.path.join(remote_basepath,month)
                                        #os.makedirs(remote_basepath,exist_ok=True)
                                        #createtd basepath 2025/jun/16-5-2025               
                                        filename = part.get_filename()
                                        filename=str(generate_random_3_digit())+filename
                                        if not filename:
                                            ext = mimetypes.guess_extension(content_type) or ".bin"
                                            filename = f"attachment_{uuid.uuid4().hex}{ext}"    
                                        filepath = os.path.join(basepath, filename)
                                        #remote_filepath=os.path.join(remote_basepath, filename)
                                        try:
                                            with open(filepath, "wb") as f:
                                                f.write(part.get_payload(decode=True))
                                            attachments.append(filepath)
                                            subject_name[filename]=msg.get("Subject", "")
                                        except Exception as save_error:
                                            print(f"Failed to save attachment {filename}: {save_error}")
                                      
                    #file scraping begins----------------------------------------------------------------------------------------------------------                     
                                        # fileorigin = os.path.basename(full_file_path) 
                                        fileorigin=filename
                                        # fileorigin_full= os.path.join(static_base, fileorigin)
                                        fileorigin_full=filepath
                                        ttt = []
                                        mmm = []
                                        aaa=[]
                                        rrr=[]
                                        invoice_format = ''
                                        wurth_start_index=False
                                        McGrath_start_index = False
                                        value = filepath
                                        yhiaustralia_start_index=False
                                        Repco_start_index=False
                                        try:
                                            with pdfplumber.open(value) as pdf:
                                                count = 0
                                                for i, page in enumerate(pdf.pages):
                                                    try:
                                                        text = page.extract_text()
                                                        if text:
                                                            # print(f"\n--- Page {i + 1} ---")
                                                            for line in text.splitlines():
                                                                if "SUA_BOS_F_DS/EUW/" in line:
                                                                    invoice_format = "WURTH"
                                                                if "McGrath Canberra Pty Ltd" in line:
                                                                    invoice_format = "John McGrath"
                                                                if "Aust Capital Terr Australia" in line:
                                                                    invoice_format = "YHI AUSTRALIA"
                                                                if "PROVIDENT MOTORS" in line:
                                                                    invoice_format = "Repco"  
                                                                count += 1
                                                                if "Delivery address Provident Motors Pty Ltd" in line:
                                                                    wurth_start_index=True
                                                                    continue
                                                                if "Your payment terms" in line:
                                                                    wurth_start_index=False
                                                                    continue
                                                                if wurth_start_index == True:
                                                                    ttt.append(line)
                                                                if "Ordered B.O. Supplied" in line:
                                                                    McGrath_start_index = True
                                                                if McGrath_start_index:
                                                                    mmm.append(line)
                                                                if "CONDITIONS OF SALE" in line:
                                                                    McGrath_start_index = False
                                                                if "S/N CODE DESCRIPTION QUANTITY UNIT PRICE AMOUNT" in line:
                                                                    yhiaustralia_start_index=True
                                                                    continue
                                                                if "SUBTOTAL" in line:
                                                                    yhiaustralia_start_index=False
                                                                    continue
                                                                if yhiaustralia_start_index ==True:
                                                                    aaa.append(line)
                                                                if "INCL GST EXCL GST TOTAL" in line:
                                                                    Repco_start_index=True
                                                                    continue
                                                                if "PAYABLE" in line:
                                                                    Repco_start_index=False

                                                                if Repco_start_index ==True:
                                                                    rrr.append(line)
                                                        else:
                                                            print(f"\n--- Page {i + 1}: No text found ---")
                                                    except Exception as e:
                                                        print(f"Error extracting text from page {i + 1}: {e}\n\n")
                                        except FileNotFoundError:
                                            print(f"The file  was not found.")
                                    
                                        if invoice_format == "WURTH":
                                            data = Invoice_Automation.convert_to_wurth_column(ttt, value, fileorigin_full, day,subject_name,fileorigin)
                                            invoicelist = cache.get('invoicelist')
                                            invoicedict=dict()
                                            invoicedict["mail_subject"]=msg.get("Subject", "")
                                            invoicedict["invoice_name"]=filename
                                            invoicedict["supplier"]="WURTH"
                                            # invoicedict["download_pdf"]=filename
                                            # invoicedict["download_excel"]=f"{filename_excel}.xlsx"
                                            invoicelist.append(invoicedict)
                                            cache.set('invoicelist', invoicelist)
                                            PurchaseReportServices.add_DataFrame_to_WurthReport(data)
                                        elif invoice_format == "John McGrath":
                                            data = Invoice_Automation.convert_to_McGrath_column(mmm, value, fileorigin_full, day,subject_name,fileorigin)
                                            invoicelist = cache.get('invoicelist')
                                            invoicedict=dict()
                                            invoicedict["mail_subject"]=msg.get("Subject", "")
                                            invoicedict["invoice_name"]=filename
                                            invoicedict["supplier"]="John McGrath"
                                            # invoicedict["download_pdf"]=filename
                                            # invoicedict["download_excel"]=f"{filename_excel}.xlsx"
                                            invoicelist.append(invoicedict)
                                            cache.set('invoicelist', invoicelist)                                          
                                            PurchaseReportServices.add_DataFrame_to_McGrathReport(data)
                                        elif invoice_format =="YHI AUSTRALIA":
                                            data=Invoice_Automation.convert_to_yhiaustralia(aaa,value,fileorigin_full,day,subject_name,fileorigin)
                                            invoicelist = cache.get('invoicelist')
                                            invoicedict=dict()
                                            invoicedict["mail_subject"]=msg.get("Subject", "")
                                            invoicedict["invoice_name"]=filename
                                            invoicedict["supplier"]="YHI AUSTRALIA"
                                            # invoicedict["download_pdf"]=filename
                                            # invoicedict["download_excel"]=f"{filename_excel}.xlsx"
                                            invoicelist.append(invoicedict)
                                            cache.set('invoicelist', invoicelist)
                                            PurchaseReportServices.add_DataFrame_to_YhiaustraliaReport(data)
                                        elif invoice_format =="Repco":
                                            data=Invoice_Automation.convert_to_repco(rrr,value,fileorigin_full,day,subject_name,fileorigin)
                                            invoicelist = cache.get('invoicelist')
                                            invoicedict=dict()
                                            invoicedict["mail_subject"]=msg.get("Subject", "")
                                            invoicedict["invoice_name"]=filename
                                            invoicedict["supplier"]="Repco"
                                            # invoicedict["download_pdf"]=filename
                                            # invoicedict["download_excel"]=f"{filename_excel}.xlsx"
                                            invoicelist.append(invoicedict)
                                            cache.set('invoicelist', invoicelist)
                                            PurchaseReportServices.add_DataFrame_to_RepcoReport(data)

                                        else:
                                            data = None
                                        
                                        PurchaseReport.objects.filter(date=day).delete()
                                        wurth_record=WurthReport.objects.filter(maildate=day).values()
                                        John_McGrath_report=McGrathReport.objects.filter(maildate=day).values()
                                        YHI_report=YhiaustraliaReport.objects.filter(maildate=day).values()
                                        Repco_report=RepcoReport.objects.filter(maildate=day).values()
                                        if Repco_report.exists():
                                            print("Repco_report exist")
                                            supplier_name="Repco"
                                            supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                                            col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                                            case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                                            col_map_list=[]
                                            for i in col_map:
                                                col_map_list.append(i)
                                            status=generatereport(Repco_report,col_map_list,case_col,supplier_name)

                                        if wurth_record.exists():
                                            print("wurth_record exist")
                                            supplier_name="wurth"
                                            supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                                            col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                                            case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                                            col_map_list=[]
                                            for i in col_map:
                                                col_map_list.append(i)
                                            status=generatereport(wurth_record,col_map_list,case_col,supplier_name)

                                        if John_McGrath_report.exists():
                                            print("John_McGrath_report exist")
                                            supplier_name="John_McGrath"
                                            supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                                            col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                                            case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                                            col_map_list=[]
                                            for i in col_map:
                                                col_map_list.append(i)
                                            status=generatereport(John_McGrath_report,col_map_list,case_col,supplier_name)
                                            print("generated:",status)
                                            
                                        if YHI_report.exists():
                                            print("YHI_report exist")
                                            supplier_name="YHI AUSTRALIA"
                                            supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                                            col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                                            case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                                            col_map_list=[]
                                            for i in col_map:
                                                col_map_list.append(i)
                                            status=generatereport(YHI_report,col_map_list,case_col,supplier_name)
                                        queryset=PurchaseReport.objects.filter(date=day).values()
                                        df = pd.DataFrame(list(queryset))
                                        df['profit'] = df['profit'].astype(str) + '%'
                                        df=df.rename(columns={
                                            'id':'S.NO','supplier':"Supplier",'date':'DATE','part_description':'PART DESCRIPTION','part_number':'PART NUMBER','trade_price':'TRADE PRICE',
                                            'total_count':'TOTAL COUNT','purchase_count':'PURCHASED COUNT','total_price':'TOTAL PRICE','actual_price':'ACTUAL PRICE',
                                            'profit':'PROFIT%','selling_price_exc_gst':'SELLING PRICE(Exc.GST)','gst':'GST','selling_price_inc_gst':'SELLING PRICE(Inc.GST)'
                                        })
                                        df['DATE'] = pd.to_datetime(df['DATE'], format='%y-%m-%d')
                                        df['DATE'] = df['DATE'].dt.strftime('%m-%d-%Y')
                                        df['TRADE PRICE'] = df['TRADE PRICE'].astype(float)
                                        df['TOTAL COUNT'] = df['TOTAL COUNT'].astype(float)
                                        df['PURCHASED COUNT'] = df['PURCHASED COUNT'].astype(float)
                                        df['TOTAL PRICE'] = df['TOTAL PRICE'].astype(float)
                                        df['ACTUAL PRICE'] = df['ACTUAL PRICE'].astype(float)
                                        df['SELLING PRICE(Exc.GST)'] = df['SELLING PRICE(Exc.GST)'].astype(float)
                                        df['GST'] = df['GST'].astype(float)
                                        df['SELLING PRICE(Inc.GST)'] = df['SELLING PRICE(Inc.GST)'].astype(float)
                                        file_excel_path=os.path.join(basepath,f"PurchaseReport_{day}.xlsx")
                                        df.to_excel(file_excel_path, index=False)
                                        template_path = os.path.join("templatefile", "templatefile.xlsx")
                                        try:
                                            wb = load_workbook(template_path, data_only=True)
                                            sheet = wb.active
                                            header_row = 4
                                            data_start_row = 5
                                            header_map = {}
                                            for col in range(1, sheet.max_column + 1):
                                                header_value = sheet.cell(row=header_row, column=col).value
                                                if header_value:
                                                    header_map[header_value.strip()] = col
            
                                            matched_columns = [col for col in df.columns if col.strip() in header_map]
                                            print("Matched columns:", matched_columns)
                                            # Write data to Excel
                                            for row_idx, row in df.iterrows():
                                                for col_name in matched_columns:
                                                    value = row[col_name]
                                                    excel_col = header_map[col_name.strip()]
                                                    sheet.cell(row=data_start_row + row_idx, column=excel_col, value=value)
                                            # Save to a dynamic name
                                            output_filename = f"PurchaseReport_{day}.xlsx"
                                            output_path = os.path.join(basepath, output_filename)
                                            wb.save(output_path)
                                            print(f"✅ Workbook saved to: {output_path}")
                                        except FileNotFoundError:
                                            print(f"❌ Template file not found: {template_path}")
                                        except Exception as e:
                                            print(f"❌ Unexpected error: {e}")
                                        # random_number = random.randint(100, 999)
                                        # file_excel_path=os.path.join(basepath,f"PurchaseReport_{day}.xlsx")
                                        # df.to_excel(file_excel_path, index=False)
                                        source_folder = rf"{basepath}"
                                        destination = rf"\\system8\D\Oakley\praveen\{basepath}"
                                        try:
                                            os.makedirs(os.path.dirname(destination), exist_ok=True)
                                            shutil.copytree(source_folder, destination, dirs_exist_ok=True)
                                            print("Folder copied successfully.")
                                        except Exception as e:
                                            print(f"Error copying folder: {e}")

                                    # elif content_type == "text/plain" and not body:
                                    #     body = part.get_payload(decode=True).decode(errors="ignore")
                                    #         else:
                                    #             body = msg.get_payload(decode=True).decode(errors="ignore")

                        emails.append({
                            "from": msg.get("From"),
                            "subject": msg.get("Subject"),
                            "date": msg.get("Date"),
                            "body": body
                        })
                    except Exception as e:
                        print(f"Error reading message {num.decode(errors='ignore')}: {e}")
                mail.logout()
        except Exception as e:
            print(f"Error connecting to Gmail: {e}")

# return True,subject_name,folder_date_path,attachments
# emails=MailAutomationClass.fetch_emails_on_date(datetime.today())
# print(emails)


class Invoice_Automation:
    @staticmethod
    def convert_to_wurth_column(ttt:List[str],filename:str,fileorigin:str,maildata:str,mail_subject:dict,fileorigin_scrap:str) -> pd.DataFrame:
        print("fileorigin:",fileorigin)
        filename=convert_to_django_media_link(filename)
        print("filename:",filename)
        print("maildata",maildata)
        supplier_name="wurth"
        itemno=[]
        item_description=[]
        customer_part_no=[]
        Ext_Net_Price_AUD=[]
        Price_Unit=[]
        Price_AUD=[]
        Quantity=[]
        Pack_Unit=[]
        ttt=[i.split() for i in ttt]
        def filter_numeric_only(data):
                pattern =  re.compile(r'^\d+(\.\d+)?$')  # Match integer or decimal numbers
                result = []
                unknown= []
                for row in data:
                    filtered_row = [item for item in row if pattern.fullmatch(item)]
                    filtered_row2 = [item for item in row if not pattern.fullmatch(item)]
                    result.append(filtered_row)
                    result.append(filtered_row2)
                return result
        result=filter_numeric_only(ttt)
        result=[i for i in result if i]
        pattern = re.compile(r'^\d+(\.\d+)?$')
        newresult=[]
        for i in range(len(result)):
            if len(result[i])==1 and pattern.fullmatch(result[i][0]):
                pass
            else:
                newresult.append(result[i])

        numberlist=[]#list contain only numbers
        finalresult=[]#list contains itemdescription and customer parts no
        for i in range(len(newresult)):
            k = True
            for j in newresult[i]:
                if not pattern.fullmatch(j):
                    k = False
            if k == True:
                if len(newresult[i]) > 6:
                    del newresult[i][0]
                    numberlist.append(newresult[i])
                    finalresult.append(newresult[i+1])
                    finalresult.append(newresult[i+2])
        item_description=[' '.join(finalresult[i]) for i in range(0,len(finalresult),2)]
        customer_part_no=[' '.join(finalresult[i]) for i in range(1,len(finalresult),2)]
        Ext_Net_Price_AUD=[i[len(i)-1] for i in numberlist]
        Price_Unit=[i[len(i)-2] for i in numberlist]
        Price_AUD=[i[len(i)-3] for i in numberlist]
        Quantity=[i[len(i)-4] for i in numberlist]
        Pack_Unit=[i[len(i)-5] for i in numberlist]
        itemno=[i[0] for i in numberlist]
        max_len = max(len(itemno),len(item_description),len(customer_part_no),len(Pack_Unit),len(Quantity),len(Price_AUD),len(Price_Unit),len(Ext_Net_Price_AUD))
        def pad(lst):
            return lst + [""] * (max_len - len(lst))
        df1=pd.DataFrame({
            "filePath":[filename for i  in range(len(pad(itemno)))],
            "supplier":[supplier_name for i in range(len(pad(itemno)))],
            "maildate":[maildata for i in range(len(pad(itemno)))],
            "itemno":pad(itemno),
            "item_description":pad(item_description),
            "customer_part_no":pad(customer_part_no),
            "Ext_Net_Price_AUD":pad(Ext_Net_Price_AUD),
            "Price_Unit":pad(Price_Unit),
            "Price_AUD":pad(Price_AUD),
            "Quantity":pad(Quantity),
            "Pack_Unit":pad(Pack_Unit)
        })

        # col_mapping=ColumnMapping.objects.filter(supplier_col=supplier_name).first()        
        # df = pd.DataFrame({
        #     "supplier":[f'{supplier_name}' for i in range(len(pad(itemno)))],
        #     "date":[f'{maildata}' for i in range(len(pad(itemno)))],
        #     "part_number": pad(globals()[col_mapping.part_number_col]),
        #     "part_description": pad(globals()[col_mapping.part_description_col]),
        #     "purchase_count":pad(globals()[col_mapping.purchase_count_col]),
        #     "trade_price":pad(globals()[col_mapping.trade_price_col]),
        #     "total_count":pad(globals()[col_mapping.total_count_col]),
        #     "total_price":pad(globals()[col_mapping.trade_price_col]),
        #     "actual_price":[globals()[col_mapping.actual_price_col] for i in range(len(pad(itemno)))],
        #     "profit":[globals()[col_mapping.profit_col] for i in range(len(pad(itemno)))],
        #     "selling_price_exc_gst":[globals()[col_mapping.selling_price_exc_gst_col] for i in range(len(pad(itemno)))],
        #     "gst":[globals()[col_mapping.gst_col] for i in range(len(pad(itemno)))],
        #     "selling_price_inc_gst":[globals()[col_mapping.selling_price_inc_gst_col] for i in range(len(pad(itemno)))]
        # })
        path_without_extension = os.path.splitext(fileorigin)[0]
        input_path_without_extension=path_without_extension.replace('out_invoice', 'in_invoice')
        df1.to_excel(f"{path_without_extension}.xlsx", index=False)
        # invoicelist = cache.get('invoicelist')
        # invoicedict=dict()
        # filename_excel=os.path.splitext(filename)[0]
        # filename_excel=filename_excel.replace('in_invoice','out_invoice')
        # invoicedict["mail_subject"]=mail_subject[fileorigin_scrap]
        # invoicedict["invoice_name"]=fileorigin_scrap
        # invoicedict["supplier"]=supplier_name
        # invoicedict["download_pdf"]=filename
        # invoicedict["download_excel"]=f"{filename_excel}.xlsx"
        # invoicelist.append(invoicedict)
        # cache.set('invoicelist', invoicelist)
        # try:
        #     Invoice=InvoiceInfo(maildate=maildata,supplier=supplier_name,ou_invoice_url=f"{path_without_extension}.xlsx",in_invoice_url=f"{input_path_without_extension}.pdf")
        #     Invoice.save()
        # except Exception as e:
        #     print(f"Error saving invoice: {e}")
        #PurchaseReportServices.add_DataFrame_to_WurthReport(df1)
        return df1
    

    @staticmethod
    def convert_to_McGrath_column(mmm : List[str],filename:str,fileorigin:str,maildata:str,mail_subject:dict,fileorigin_scrap:str) -> pd.DataFrame:
        filename=convert_to_django_media_link(filename)
        print("maildate:",maildata)
        supplier_name="John_McGrath"
        location= []
        part_Number= []
        description= []
        ordered= []
        supplied= []
        unit_List= []
        unit_Net= []
        GST_Code= []
        total= []
        def filter_numeric_only(data):
            pattern = re.compile(r'^\d+(\.\d+)?$')  # Match integer or decimal numbers
            result = []
            unknown= []
            for row in data:
                filtered_row = [item for item in row if pattern.fullmatch(item)]
                filtered_row2 = [item for item in row if not pattern.fullmatch(item)]
                result.append(filtered_row)
                unknown.append(filtered_row2)
            return result,unknown
        mmm=mmm[1:len(mmm)-1]
        mmm=[i.split() for i in mmm]
        classified_data,unknown = filter_numeric_only(mmm)
        classified_data2= []
        for i in classified_data:
            if len(i)==7:
                del i[3]
            if not i:
                i=[0,0,0,0,0,0]
            classified_data2.append(i)
        classified_data=classified_data2
        classified_data=[ i[1:] for i in classified_data]
        unknown=[i for i in unknown]
        ordered=[i[0] for i in classified_data]
        supplied=[i[1] for i in classified_data]
        unit_List= [i[2] for i in classified_data]
        unit_Net=[i[3] for i in classified_data]
        total=[i[4] for i in classified_data]
        location=[i[0] for i in unknown]
        part_Number=[i[1] for i in unknown]
        GST_Code=[i[len(i)-1] for i in unknown]
        description=[' '.join(i[2:len(i)-1]) for i in unknown]
        max_len = max(len(location),len(part_Number),len(description),len(ordered),len(supplied),len(unit_List),len(unit_Net),len(GST_Code),len(total))
        def pad(lst):
            return lst + [""] * (max_len - len(lst))
        
        df1=pd.DataFrame({
            "filePath":[f'{filename}' for i  in range(len(pad(location)))],
            "supplier":[f'{supplier_name}' for i in range(len(pad(location)))],
            "maildate":[maildata for i in range(len(pad(location)))],
            "location": pad(location),
            "part_Number": pad(part_Number),
            "description":pad(description),
            "ordered":pad(ordered),
            "supplied":pad(supplied),
            "unit_List":pad(unit_List),
            "unit_Net":pad(unit_Net),
            "GST_Code":pad(GST_Code),
            "total":pad(total)
        })

        # col_mapping=ColumnMapping.objects.filter(supplier_col=supplier_name).first() 
        # df = pd.DataFrame({
        #     "supplier":[f'{supplier_name}' for i in range(len(pad(location)))],
        #     "date":[f'{maildata}' for i in range(len(pad(location)))],
        #     "part_number":pad(globals()[col_mapping.part_number_col]),
        #     "part_description":pad(globals()[col_mapping.part_description_col]),
        #     "purchase_count":pad(globals()[col_mapping.purchase_count_col]),
        #     "trade_price":pad(globals()[col_mapping.trade_price_col]),
        #     "total_count":[globals()[col_mapping.total_count_col] for i in range(len(pad(location)))],
        #     "total_price":pad(globals()[col_mapping.total_price_col]),
        #     "actual_price":[globals()[col_mapping.actual_price_col] for i in range(len(pad(part_Number)))],
        #     "profit":[globals()[col_mapping.profit_col] for i in range(len(pad(part_Number)))],
        #     "selling_price_exc_gst":[globals()[col_mapping.selling_price_exc_gst_col] for i in range(len(pad(part_Number)))],
        #     "gst":[globals()[col_mapping.gst_col] for i in range(len(pad(part_Number)))],
        #     "selling_price_inc_gst":[globals()[col_mapping.selling_price_inc_gst_col] for i in range(len(pad(part_Number)))]
        # })
        #print(df1)

        path_without_extension = os.path.splitext(fileorigin)[0]
        input_path_without_extension=path_without_extension.replace('out_invoice', 'in_invoice')
        df1.to_excel(f"{path_without_extension}.xlsx", index=False)

        # invoicelist = cache.get('invoicelist')
        # invoicedict=dict()
        # filename_excel=os.path.splitext(filename)[0]
        # filename_excel=filename_excel.replace('in_invoice','out_invoice')
        # invoicedict["mail_subject"]=mail_subject[fileorigin_scrap]
        # invoicedict["invoice_name"]=fileorigin_scrap
        # invoicedict["supplier"]=supplier_name
        # invoicedict["download_pdf"]=filename
        # invoicedict["download_excel"]=f"{filename_excel}.xlsx"
        # invoicelist.append(invoicedict)
        # cache.set('invoicelist', invoicelist)


        # try:
        #     Invoice=InvoiceInfo(maildate=maildata,ou_invoice_url=f"{path_without_extension}.xlsx",in_invoice_url=f"{input_path_without_extension}.pdf")
        #     Invoice.save()
        # except Exception as e:
        #     print(f"Error saving invoice: {e}")
        #PurchaseReportServices.add_DataFrame_to_McGrathReport(df1)
        return df1
    
    @staticmethod
    def convert_to_yhiaustralia(aaa:List[str],filename:str,fileorigin:str,maildata:str,mail_subject:dict,fileorigin_scrap:str) -> pd.DataFrame:
        filename=convert_to_django_media_link(filename)
        supplier_name="YHI AUSTRALIA"
        CODE= []
        DESCRIPTION= []
        QUANTITY= []
        UNIT_PRICE= []
        AMOUNT= []
        aaa=[i.split() for i in aaa]
        aaa_dict_list=[]
        extra=0
        pos=[]
        for i in range(1,len(aaa)):
            if len(aaa[i])< 7:
                aaa[i-1].insert(2,''.join(aaa[i]))
                pos.append(i)

        for i in pos:
            if i in aaa:
                aaa.remove(i)
        aaa = [i for i in aaa if len(i) >= 6]
                
        
        def filter_numeric_only(data):
                pattern = re.compile(r'^\d+(\.\d+)?$')  # Match integer or decimal numbers
                result = []
                unknown= []
                for row in data:
                    filtered_row = [item for item in row if pattern.fullmatch(item)]
                    filtered_row2 = [item for item in row if not pattern.fullmatch(item)]
                    result.append(filtered_row)
                    unknown.append(filtered_row2)
                return result,unknown
        
        numbers,texts=filter_numeric_only(aaa)
        CODE=[i[0] for i in texts]
        DESCRIPTION=[' '.join(i[1:]) for i in texts]
        AMOUNT=[i[3] for i in numbers]
        UNIT_PRICE=[i[2] for i in numbers]
        QUANTITY=[i[1] for i in numbers]
        max_len = max(len(CODE),len(DESCRIPTION),len(AMOUNT),len(UNIT_PRICE),len(QUANTITY))
        def pad(lst):
            return lst + [""] * (max_len - len(lst))
        
        df1=pd.DataFrame({
                        "filePath":[f'{filename}' for i  in range(len(pad(CODE)))],
                        "supplier":[f'{supplier_name}' for i in range(len(pad(CODE)))],
                        "maildate":[f'{maildata}' for i in range(len(pad(CODE)))],          
                        "code":pad(CODE),
                        "description":pad(DESCRIPTION),
                        "quantity":pad(QUANTITY),
                        "unit_price":pad(UNIT_PRICE),
                        "amount":pad(AMOUNT)
                        })
        
        # col_mapping=ColumnMapping.objects.filter(supplier_col=supplier_name).first()
        # df=pd.DataFrame({"supplier":[f'{supplier_name}' for i in range(len(pad(CODE)))],
        #                  "date":[f'{maildata}' for i in range(len(pad(CODE)))],
        #                 "part_number":pad(globals()[col_mapping.part_number_col]),
        #                 "part_description":pad(globals()[col_mapping.part_description_col]),
        #                 "purchase_count":pad(globals()[col_mapping.purchase_count_col]),
        #                 "trade_price":pad(globals()[col_mapping.trade_price_col]),
        #                 "total_count":[globals()[col_mapping.total_count_col] for i in range(len(pad(CODE)))],
        #                 "total_price":pad(globals()[col_mapping.total_price_col]),
        #                 "actual_price":[globals()[col_mapping.actual_price_col] for i in range(len(pad(CODE)))],
        #                 "profit":[globals()[col_mapping.profit_col] for i in range(len(pad(CODE)))],
        #                 "selling_price_exc_gst":[globals()[col_mapping.selling_price_exc_gst_col] for i in range(len(pad(CODE)))],
        #                 "gst":[globals()[col_mapping.gst_col] for i in range(len(pad(CODE)))],
        #                 "selling_price_inc_gst":[globals()[col_mapping.selling_price_inc_gst_col] for i in range(len(pad(CODE)))]
        #                 })
     
        path_without_extension = os.path.splitext(fileorigin)[0]
        input_path_without_extension=path_without_extension.replace('out_invoice', 'in_invoice')
        df1.to_excel(f"{path_without_extension}.xlsx", index=False)
        # invoicelist = cache.get('invoicelist')
        # invoicedict=dict()
        # filename_excel=os.path.splitext(filename)[0]
        # filename_excel=filename_excel.replace('in_invoice','out_invoice')
        # invoicedict["mail_subject"]=mail_subject[fileorigin_scrap]
        # invoicedict["invoice_name"]=fileorigin_scrap
        # invoicedict["supplier"]=supplier_name
        # invoicedict["download_pdf"]=filename
        # invoicedict["download_excel"]=f"{filename_excel}.xlsx"
        # invoicelist.append(invoicedict)
        # cache.set('invoicelist', invoicelist)
        # try:
        #     Invoice=InvoiceInfo(maildate=maildata,ou_invoice_url=f"{path_without_extension}.xlsx",in_invoice_url=f"{input_path_without_extension}.pdf")
        #     Invoice.save()
        # except Exception as e:
        #     print(f"Error saving invoice: {e}")
        # PurchaseReportServices.add_DataFrame_to_YhiaustraliaReport(df1)
        return df1
    
    @staticmethod
    def convert_to_repco(rrr:list[str],filename:str,fileorigin:str,maildata:str,mail_subject:dict,fileorigin_scrap:str):
        filename=convert_to_django_media_link(filename)
        supplier_name="Repco"
        rrr=rrr[:len(rrr)-1]
        rrr=[i.split() for i in rrr]
        rrr=[i[1:] for i in rrr]
        part_number=[]
        description=[]
        retail_incl_gst=[]
        uom=[]
        # qty_ordered=[]
        # back_ordered=[]
        qty_supplied=[]
        unit_price_excl_gst=[]
        s=[]
        total_gst=[]
        total_incl_gst=[]
        numbers=[]
        for r in rrr:
            found=[re.findall(r'\d+\.?\d*', i) for i in r]
            found=[i for i in found if i]
            numbers.append(found)
        numeric_list = []
        alphabet_list = []
        for row in rrr:
            row_numeric = []
            row_alpha = []
            for item in row:
                clean_item = item.replace(',', '')
                try:
                    row_numeric.append(float(clean_item))
                except ValueError:
                    row_alpha.append(item)
            numeric_list.append(row_numeric)
            alphabet_list.append(row_alpha)
        alphabet_list=[i[:len(i)-1] if "S" in i else i for i in alphabet_list]        
        print("Numeric List:", numeric_list)
        # alphabet_list_trim = []
        # for row in alphabet_list:
        #     if 'EACH' in row:
        #         idx = row.index('EACH')
        #         alphabet_list_trim.append(row[:idx + 1]) 
        #     else:
        #         alphabet_list_trim.append(row)
        # alphabet_list=alphabet_list_trim
        print("Alphabet List:", alphabet_list)
        total_incl_gst=[i[len(i)-1] if len(i) >= 1 else None for i in numeric_list]
        total_gst=[i[len(i)-2] if len(i) >= 2 else None for i in numeric_list]
        s=[i[len(i)-3] if len(i) >= 3 else None for i in numeric_list]
        unit_price_excl_gst=[i[len(i)-4] if len(i) >= 4 else None for i in numeric_list]
        qty_supplied=[i[len(i)-5] if len(i) >= 5 else None for i in numeric_list]
        retail_incl_gst=[i[len(i)-6] if len(i) >= 6 else None for i in numeric_list]
        part_number = [i.pop(0) for i in alphabet_list]
        uom=[i.pop(len(i)-1) for i in alphabet_list]
        description=[' '.join(i) for i in alphabet_list]
        print("total_incl_gst",total_incl_gst)
        print("total_gst",total_gst)
        print("s",s)
        print("unit_price_excl_gst",unit_price_excl_gst)
        print("qty_supplied",qty_supplied)
        print("retail_incl_gst",retail_incl_gst)
        print("part_number",part_number)
        print("uom",uom)
        print("description",description)
        max_len = max(len(total_incl_gst),len(total_gst),len(unit_price_excl_gst),len(qty_supplied),len(part_number))
        def pad(lst):
            return lst + [""] * (max_len - len(lst))
        df1=pd.DataFrame({
                            # "filePath":[f'{filename}' for i  in range(len(pad(part_number)))],
                            # "supplier":[f'{supplier_name}' for i in range(len(pad(part_number)))],
                            # "maildate":[f'{maildata}' for i in range(len(pad(part_number)))],
                            "part_number":pad(part_number),
                            "description":pad(description),
                            "uom":pad(uom),
                            "retail_incl_gst":pad(retail_incl_gst),
                            "unit_price_excl_gst":pad(unit_price_excl_gst),
                            "qty_supplied":pad(qty_supplied),
                            "total_gst":pad(total_gst),
                            "s":pad(s),
                            "total_incl_gst":pad(total_incl_gst)
                            })
        path_without_extension = os.path.splitext(fileorigin)[0]
        input_path_without_extension=path_without_extension.replace('out_invoice', 'in_invoice')
        invoicelist = cache.get('invoicelist')
        invoicedict=dict()
        filename_excel=os.path.splitext(filename)[0]
        filename_excel=filename_excel.replace('in_invoice','out_invoice')
        invoicedict["mail_subject"]=mail_subject[fileorigin_scrap]
        invoicedict["invoice_name"]=fileorigin_scrap
        invoicedict["supplier"]=supplier_name
        invoicedict["download_pdf"]=filename
        invoicedict["download_excel"]=f"{filename_excel}.xlsx"
        invoicelist.append(invoicedict)
        cache.set('invoicelist', invoicelist)
        df1.to_excel(f"{path_without_extension}.xlsx", index=False)
        return df1

class UtilityClasses:
    @staticmethod
    def scrap(full_file_path, maildata,mail_subject):
        static_base = os.path.join(settings.OUTPUT_INVOICE,maildata)
        if not os.path.exists(static_base):
            os.makedirs(static_base)
        fileorigin = os.path.basename(full_file_path)
        fileorigin_full= os.path.join(static_base, fileorigin)
        ttt = []
        mmm = []
        aaa=[]
        rrr=[]
        invoice_format = ''
        wurth_start_index=False
        McGrath_start_index = False
        value = full_file_path
        yhiaustralia_start_index=False
        Repco_start_index=False

        try:
            with pdfplumber.open(value) as pdf:
                count = 0
                for i, page in enumerate(pdf.pages):
                    try:
                        text = page.extract_text()
                        if text:
                            # print(f"\n--- Page {i + 1} ---")
                            for line in text.splitlines():
                                if "SUA_BOS_F_DS/EUW/" in line:
                                    invoice_format = "WURTH"
                                if "McGrath Canberra Pty Ltd" in line:
                                    invoice_format = "John McGrath"
                                if "Aust Capital Terr Australia" in line:
                                    invoice_format = "YHI AUSTRALIA"
                                if "PROVIDENT MOTORS" in line:
                                    invoice_format = "Repco"   
                                count += 1
                                if "Delivery address Provident Motors Pty Ltd" in line:
                                    wurth_start_index=True
                                    continue
                                if "Your payment terms" in line:
                                    wurth_start_index=False
                                    continue
                                if wurth_start_index == True:
                                    ttt.append(line)
                                if "Ordered B.O. Supplied" in line:
                                    McGrath_start_index = True
                                if McGrath_start_index:
                                    mmm.append(line)
                                if "CONDITIONS OF SALE" in line:
                                    McGrath_start_index = False
                                if "S/N CODE DESCRIPTION QUANTITY UNIT PRICE AMOUNT" in line:
                                    yhiaustralia_start_index=True
                                    continue
                                if "SUBTOTAL" in line:
                                    yhiaustralia_start_index=False
                                    continue
                                if yhiaustralia_start_index ==True:
                                    aaa.append(line)
                                                                
                                if "INCL GST EXCL GST TOTAL" in line:
                                    Repco_start_index=True
                                    continue
                                    
                                if "PAYABLE" in line:
                                    Repco_start_index=False

                                if Repco_start_index ==True:
                                    rrr.append(line)
                        else:
                            print(f"\n--- Page {i + 1}: No text found ---")
                    except Exception as e:
                        print(f"Error extracting text from page {i + 1}: {e}\n\n")
        except FileNotFoundError:
            print(f"The file  was not found.")
            print("invoice format:", invoice_format)
        if invoice_format == "WURTH":
            data = Invoice_Automation.convert_to_wurth_column(ttt, value, fileorigin_full, maildata,mail_subject,fileorigin)
        elif invoice_format == "John McGrath":
            data = Invoice_Automation.convert_to_McGrath_column(mmm, value, fileorigin_full, maildata,mail_subject,fileorigin)
        elif invoice_format =="YHI AUSTRALIA":
            data=Invoice_Automation.convert_to_yhiaustralia(aaa,value,fileorigin_full,maildata,mail_subject,fileorigin)
        elif invoice_format == "Repco":
            data=Invoice_Automation.convert_to_repco(rrr,value,fileorigin_full,maildata,mail_subject,fileorigin)
        else:
            data = None
        return data
    


#scrapping for testing purpose------------------------------------------------
def scrap_for_test(full_file_path):
    ttt = []#SET FOR WURTH INVOICE SCRAP
    mmm = []#SET FOR MCGRATH INVOICE SCRAP
    aaa=[]#SET FOR YHI INVOICE SCRAP
    invoice_format = ''
    wurth_start_index=False
    McGrath_start_index = False
    value = full_file_path
    yhiaustralia_start_index=False
    try:
        with pdfplumber.open(value) as pdf:
            count = 0
            for i, page in enumerate(pdf.pages):
                try:
                    text = page.extract_text()
                    if text:
                        #print(f"\n--- Page {i + 1} ---")
                        for line in text.splitlines():
                            if "SUA_BOS_F_DS/EUW/" in line:
                                invoice_format = "WURTH"
                            if "McGrath Canberra Pty Ltd" in line:
                                invoice_format = "John McGrath"
                            if "Aust Capital Terr Australia" in line:
                                invoice_format = "YHI AUSTRALIA"
                            count += 1
            
                            #print(f"{count}-----{line}--------")
                            if "Delivery address Provident Motors Pty Ltd" in line:
                                wurth_start_index=True
                                continue

                            if "Your payment terms" in line:
                                wurth_start_index=False
                                continue

                            if wurth_start_index == True:
                                ttt.append(line)
                              
                            if "Ordered B.O. Supplied" in line:
                                McGrath_start_index = True

                            if McGrath_start_index == True:
                                mmm.append(line)

                            if "CONDITIONS OF SALE" in line:
                                McGrath_start_index = False

                            if "S/N CODE DESCRIPTION QUANTITY UNIT PRICE AMOUNT" in line:
                                yhiaustralia_start_index=True
                                continue
                            
                            if "SUBTOTAL" in line:
                                yhiaustralia_start_index=False
                                continue
                            
                            if yhiaustralia_start_index ==True:
                                aaa.append(line)
                    else:
                        pass
                        #print(f"\n--- Page {i + 1}: No text found ---")
                except Exception as e:
                        print(f"Error extracting text from page {i + 1}: {e}\n\n")

    except FileNotFoundError:
        print(f"The file  was not found.")
    print("invoice supplier:",invoice_format)
    if invoice_format == "WURTH":
        Invoice_Automation.convert_to_wurth_column(ttt,value)
    elif invoice_format == "John McGrath":
        data = Invoice_Automation.convert_to_McGrath_column(mmm)
    elif invoice_format =="YHI AUSTRALIA":
        Invoice_Automation.convert_to_yhiaustralia(aaa)
    else:
        data = None

def convert_to_django_media_link(windows_path):
    parts = windows_path.split(os.sep)
    try:
        media_index = parts.index('media')
        relative_path_parts = parts[media_index:]
        django_path = '/'.join(relative_path_parts)
        return f'/media/{django_path[len("media/"):]}' # Remove the initial "media/" if it exists
    except ValueError:
        return "Error: 'media' directory not found in the path."
    

#----------------------------------------------------------generate report functions--------------------------------------------------------------------------
from .models import PurchaseReport

def profit_fun(casedict):
    profit=casedict['profit']
    try:
        profit=float(profit.replace('%', '').replace(' ', ''))
    except ValueError:
        profit= 0.0
    return profit

def selling_price_exc_gst_fun(profit,actualprice):
    profit=(actualprice/100)*profit
    selling_price_exc_gst = actualprice+profit
    return selling_price_exc_gst

def gst_fun(casedict,selling_price_exc_gst):
    gst=casedict['gst']
    try:
        gst = float(gst.replace('%', '').replace(' ', ''))
    except ValueError:
        gst = 0.0

    gstvalue=(selling_price_exc_gst/100)*gst
    return gstvalue

def selling_price_inc_gst_fun(gstvalue,selling_price_exc_gst):
    selling_price_inc_gst=selling_price_exc_gst+gstvalue
    return selling_price_inc_gst

def checkcase(p_report,cases):
    for i in cases:
        if p_report['actual_price'] >= i['minvalue'] and p_report['actual_price'] <= i['maxvalue']:
            return i

def wurth_actucalprice(report_i):
    netprice_aud=report_i['Ext_Net_Price_AUD']
    try:
        netprice_aud = float(netprice_aud.replace('%', '').replace(' ', ''))
    except ValueError:
        netprice_aud=0.0

    quantity=report_i['Quantity']
    try:
        quantity =  float(quantity.replace('%', '').replace(' ', ''))
    except ValueError:
        quantity=0.0
    report_i['actual_price']=netprice_aud/quantity
    return report_i


def johnmcgrath_actualprice(report_i):
    unitnet=report_i['unit_Net']
    try:
        unitnet=float(unitnet.replace('%', '').replace(' ', ''))
    except ValueError:
        unitnet=0.0
    gstcode=report_i['GST_Code']
    try:
        gstcode=float(gstcode.replace('%', '').replace(' ', ''))
    except ValueError:
        gstcode=0.0
    gstcode=(unitnet/100)*gstcode
    report_i['actual_price']=unitnet+gstcode
    return report_i


def yhi_actualprice(report_i):
    unitprice=report_i['unit_price']
    try:
        unitprice=float(unitprice.replace('%', '').replace(' ', ''))
    except ValueError:
        unitprice=0.0
    report_i['actual_price']=unitprice
    return report_i

def repco_actualprice(report_i):
    retail_incl_gst=report_i['retail_incl_gst']
    print("retail_incl_gst:",retail_incl_gst)
    try:
        retail_incl_gst=float(retail_incl_gst.replace('%', '').replace(' ', ''))
        actual_price=float(retail_incl_gst-(retail_incl_gst*0.0909))
    except ValueError:
        actual_price=0.0
        print("report_i:",report_i)
    report_i['actual_price']=actual_price
    return report_i
    


def generatereport(report,col_map_list,case_col,supplier_name):
    if len(case_col) >=1:
        col_map_list=col_map_list[0]
        col_map_list.pop('id')
        for i in report:
            do_calculation(i,col_map_list,case_col,supplier_name)
    return True
                

def do_calculation(report_i,col_map,cases,supplier_name): 
    report_i.pop('id')
    report_i.pop('filePath')
    if supplier_name == "wurth":
        report_i= wurth_actucalprice(report_i)
    elif supplier_name == "YHI AUSTRALIA":
        report_i=yhi_actualprice(report_i)
    elif supplier_name == "John_McGrath":
        report_i=johnmcgrath_actualprice(report_i)
    elif supplier_name == "Repco":
        report_i=repco_actualprice(report_i)
    
    report_i_list=list(report_i.keys())
    p_report=dict()
    for i,j in col_map.items():
        if j in report_i_list:
            p_report[i.strip('_col')]=report_i[j]
    p_report['actual_price']=round(report_i['actual_price'],2)
    casedict=checkcase(p_report,cases)
    actual_price=p_report['actual_price']
    profit=int(profit_fun(casedict))
    selling_price_exc_gst=round(selling_price_exc_gst_fun(profit,actual_price),2)
    gst=round(gst_fun(casedict,selling_price_exc_gst),2)
    selling_price_inc_gst=round(selling_price_inc_gst_fun(gst,selling_price_exc_gst),2)
    p_report['profit']=profit
    p_report['selling_price_exc_gst']=selling_price_exc_gst
    p_report['gst']=gst
    p_report['selling_price_inc_gst']=selling_price_inc_gst
    try:
        p_object=PurchaseReport.objects.create(supplier=supplier_name,date=p_report['date'],part_description=p_report['part_description'],part_number=p_report['part_number'],trade_price=p_report['trade_price'],total_count=p_report['total_count'],purchase_count=p_report['purchase_count'],
                                            total_price=p_report['total_price'],actual_price=p_report['actual_price'],profit=p_report['profit'],selling_price_exc_gst=p_report['selling_price_exc_gst'],gst=p_report['gst'],selling_price_inc_gst=p_report['selling_price_inc_gst'])
        p_object.save()
    except Exception as e:
        print("Unexpected error:", str(e))
    print("saved")


def generate_report_to_xlsl(record,inputDate):
        print("maildate",inputDate)
        filename= f"purchase_report_{inputDate}.xlsx"
        df = pd.DataFrame({
            "supplier":[i["supplier"] for i in record],
            "date":[i['date'] for i in record],
            "part_number":[i['part_number'] for i in record],
            "part_description":[i['part_description'] for i in record],
            "purchase_count":[i['purchase_count'] for i in record],
            "trade_price":[i['trade_price'] for i in record],
            "total_count":[i['total_count'] for i in record],
            "total_price":[i['trade_price'] for i in record],
            "actual_price":[i['actual_price'] for i in record],
            "profit":[i['profit'] for i in record],
            "selling_price_exc_gst":[i['selling_price_exc_gst'] for i in record],
            "gst":[i['gst'] for i in record],
            "selling_price_inc_gst":[i['selling_price_inc_gst'] for i in record]
        })
        df.to_excel(filename, index=False)

def date_extraction():
    maildate=datetime.today()
    date_str =  maildate.strftime("%d-%m-%y")
    date_obj = datetime.strptime(date_str, "%d-%m-%y")
    year = date_obj.year
    month = date_obj.strftime("%b")
    day = date_obj.day
    print(year, month, day)


