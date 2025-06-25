from django.shortcuts import render
from django.http import HttpResponse,Http404,FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from django.core.cache import cache
from django.http import JsonResponse
import pdfplumber
import pandas as pd
import re
import json 
from .utils import Invoice_Automation,MailAutomationClass,UtilityClasses
from datetime import datetime,date
from django.conf import settings
import os
from .models import InvoiceInfo,PurchaseReport,Supplier,ColumnMapping,CaseModel,ScrappedDate
from .models import WurthReport,McGrathReport,YhiaustraliaReport,RepcoReport
from .service_repository import PurchaseReportServices
from main import service_repository
from .utils import generatereport,generate_report_to_xlsl
import random
import shutil
import stat
import zipfile
from openpyxl import load_workbook
import glob

def remove_readonly(func, path, _):
    os.chmod(path, stat.S_IWRITE)
    func(path)

def home(request):
    return render(request,'index2.html')

def reporview(request):
    return render(request,"reportview.html")

def generated_report(request):
    return render(request,"generated_report.html")



@csrf_exempt
def uploadpdf_invoice(request):
    if request.method == 'POST':
        files = request.FILES.getlist('files')
        for f in files:
            print("filename:",f.name)
            # with open(file_path, 'wb+') as destination:
            #     for chunk in f.chunks():
            #         destination.write(chunk)
        return JsonResponse({'status': 'success'})

@csrf_exempt
def showunseen_mails(request):
    if request.method == 'GET':
        try:
            data=invoice_list=cache.get('invoicelist')
        # return JsonResponse({'status': 'success', 'invoice_list': invoice_list},safe=False)
            return JsonResponse({'invoice_list': data})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def filecorrection(request):
    if request.method == 'POST':
        files = request.FILES.getlist('files')
        correction=request.POST.get('corr_date')
        correction = datetime.strptime(correction, "%Y-%m-%d").date()
        date_str =  correction.strftime("%y-%m-%d")
        date_obj = datetime.strptime(date_str, "%y-%m-%d")
        year = str(date_obj.year)
        month = date_obj.strftime("%b")
        day = date_str
        print("year:",year)
        print("month:",month)
        print("day:",day)
        print("correction date:",correction)
        excel_files=[]
        if correction:
            if files:
                target_folder =  os.path.join('media',str(year),str(month),str(day))
                print("target folder:",target_folder)

                print("target_folder",target_folder)
                os.makedirs(target_folder, exist_ok=True)
                for f in files:
                    if f.name.endswith('.xlsx'):
                        path = os.path.join(target_folder, f.name)
                        excel_files.append(path)
                        print("filepath1:",path)
                        if os.path.exists(path):
                            os.remove(path)
                        with open(path, 'wb+') as destination:
                            for chunk in f.chunks():
                                destination.write(chunk)
                #excel_files = glob.glob(os.path.join(target_folder, "*.xlsx"))
                for f in excel_files:
                    print("list all files:",f)
                    sheets = pd.read_excel(f, engine='openpyxl',dtype=str)
                    if sheets['supplier'].iloc[0]=="wurth":
                        WurthReport.objects.filter(maildate=day).delete()
                        PurchaseReport.objects.filter(date=day,supplier="wurth").delete()
                        PurchaseReportServices.add_DataFrame_to_WurthReport(sheets)
                    elif sheets['supplier'].iloc[0]=="John_McGrath":
                        McGrathReport.objects.filter(maildate=day).delete()
                        PurchaseReport.objects.filter(date=day,supplier="John_McGrath").delete()
                        PurchaseReportServices.add_DataFrame_to_McGrathReport(sheets)
                    elif sheets['supplier'].iloc[0]=="YHI AUSTRALIA":
                        YhiaustraliaReport.objects.filter(maildate=day).delete()
                        PurchaseReport.objects.filter(date=day,supplier="YHI AUSTRALIA").delete()
                        PurchaseReportServices.add_DataFrame_to_YhiaustraliaReport(sheets)
                    elif sheets['supplier'].iloc[0]=="Repco":
                        RepcoReport.objects.filter(maildate=day).delete()
                        PurchaseReport.objects.filter(date=day, supplier="Repco").delete()
                        PurchaseReportServices.add_DataFrame_to_RepcoReport(sheets)
                    wurth_record=WurthReport.objects.filter(maildate=day).values()
                    John_McGrath_report=McGrathReport.objects.filter(maildate=day).values()
                    YHI_report=YhiaustraliaReport.objects.filter(maildate=day).values()
                    Repco_report=RepcoReport.objects.filter(maildate=day).values()

                    if Repco_report.exists():
                        print("Repco_report exist")
                        supplier_name="Repco"
                        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                        col_map_list=[]
                        for i in col_map:
                            col_map_list.append(i)
                        status=generatereport(Repco_report,col_map_list,case_col,supplier_name)
                    if wurth_record.exists():
                        print("wurth_record exist")
                        supplier_name="wurth"
                        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                        col_map_list=[]
                        for i in col_map:
                            col_map_list.append(i)
                        status=generatereport(wurth_record,col_map_list,case_col,supplier_name)
                    if John_McGrath_report.exists():
                        print("John_McGrath_report exist")
                        supplier_name="John_McGrath"
                        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                        col_map_list=[]
                        for i in col_map:
                            col_map_list.append(i)
                        status=generatereport(John_McGrath_report,col_map_list,case_col,supplier_name)
                        print("generated:",status)
                    if YHI_report.exists():
                        print("YHI_report exist")
                        supplier_name="YHI AUSTRALIA"
                        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
                        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
                        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
                        col_map_list=[]
                        for i in col_map:
                            col_map_list.append(i)
                        status=generatereport(YHI_report,col_map_list,case_col,supplier_name)
            queryset=PurchaseReport.objects.filter(date=day).values()
            df = pd.DataFrame(list(queryset))
            print(df)
            df['profit'] = df['profit'].astype(str) + '%'
            df=df.rename(columns={
                'id':'S.NO','supplier':"Supplier",'date':'DATE','part_description':'PART DESCRIPTION','part_number':'PART NUMBER','trade_price':'TRADE PRICE',
                'total_count':'TOTAL COUNT','purchase_count':'PURCHASED COUNT','total_price':'TOTAL PRICE','actual_price':'ACTUAL PRICE',
                'profit':'PROFIT%','selling_price_exc_gst':'SELLING PRICE(Exc.GST)','gst':'GST','selling_price_inc_gst':'SELLING PRICE(Inc.GST)'
            })
            df['DATE'] = pd.to_datetime(df['DATE'], format='%y-%m-%d')
            df['DATE'] = df['DATE'].dt.strftime('%m-%d-%Y')
            df['TRADE PRICE'] = df['TRADE PRICE'].astype(float)
            df['TOTAL COUNT'] = df['TOTAL COUNT'].astype(float)
            df['PURCHASED COUNT'] = df['PURCHASED COUNT'].astype(float)
            df['TOTAL PRICE'] = df['TOTAL PRICE'].astype(float)
            df['ACTUAL PRICE'] = df['ACTUAL PRICE'].astype(float)
            df['SELLING PRICE(Exc.GST)'] = df['SELLING PRICE(Exc.GST)'].astype(float)
            df['GST'] = df['GST'].astype(float)
            df['SELLING PRICE(Inc.GST)'] = df['SELLING PRICE(Inc.GST)'].astype(float)
            file_excel_path=os.path.join(target_folder,f"PurchaseReport_{day}.xlsx")
            df.to_excel(file_excel_path, index=False)
            template_path = os.path.join("templatefile", "templatefile.xlsx")
            try:
                wb = load_workbook(template_path, data_only=True)
                sheet = wb.active
                header_row = 4
                data_start_row = 5
                header_map = {}
                for col in range(1, sheet.max_column + 1):
                    header_value = sheet.cell(row=header_row, column=col).value
                    if header_value:
                        header_map[header_value.strip()] = col

                # ✅ Only keep columns from DataFrame that match Excel headers
                matched_columns = [col for col in df.columns if col.strip() in header_map]
                print("Matched columns:", matched_columns)
                # Write data to Excel
                for row_idx, row in df.iterrows():
                    for col_name in matched_columns:
                        value = row[col_name]
                        excel_col = header_map[col_name.strip()]
                        sheet.cell(row=data_start_row + row_idx, column=excel_col, value=value)
                # Save to a dynamic name
                output_filename = f"PurchaseReport_{day}.xlsx"
                output_path = os.path.join(target_folder, output_filename)
                wb.save(output_path)
                print(f"✅ Workbook saved to: {output_path}")
            except FileNotFoundError:
                print(f"❌ Template file not found: {template_path}")
            except Exception as e:
                print(f"❌ Unexpected error: {e}")
            
            source_folder = rf"{target_folder}"
            destination = rf"\\system8\D\Oakley\praveen\{target_folder}"
            try:
                os.makedirs(os.path.dirname(destination), exist_ok=True)
                shutil.copytree(source_folder, destination, dirs_exist_ok=True)
                print("Folder copied successfully.")
            except Exception as e:
                print(f"Error copying folder: {e}")
        return JsonResponse({"status": "success",'message': 'Files converted successfully'})
    else:
        return JsonResponse({'error': "Failed to connect"}, status=500)

    


@csrf_exempt
def zip_and_download(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        maildate = data.get('maildate')
        target_out_folder = os.path.join(settings.MEDIA_ROOT, "out_invoice", maildate)
        output_zip = f'media/{maildate}.zip'
        with zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(target_out_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Add file with relative path to preserve folder structure
                    arcname = os.path.relpath(file_path, target_out_folder)
                    zipf.write(file_path, arcname)
        response = FileResponse(open(output_zip, 'rb'), as_attachment=True, filename=f"{maildate}.zip")
        return response
    

@csrf_exempt
def delete_mailed_records(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            maildate = data.get('maildate')
            target_in_folder = os.path.join(settings.MEDIA_ROOT, "in_invoice", maildate)
            target_out_folder = os.path.join(settings.MEDIA_ROOT, "out_invoice", maildate)
            for folder in [target_in_folder, target_out_folder]:
                try:
                    if os.path.exists(folder):
                        shutil.rmtree(folder, onerror=remove_readonly)
                        print(f"Deleted: {folder}")
                    else:
                        print(f"Folder not found: {folder}")
                except Exception as folder_err:
                    print(f"Error deleting folder {folder}: {folder_err}")
                    return JsonResponse({'error': f"Failed to delete folder: {folder}, {str(folder_err)}"}, status=500)
            WurthReport.objects.filter(maildate=maildate).delete()
            McGrathReport.objects.filter(maildate=maildate).delete()
            YhiaustraliaReport.objects.filter(maildate=maildate).delete()
            PurchaseReport.objects.filter(date=maildate).delete()
            ScrappedDate.objects.filter(maildate=maildate).delete()

            return JsonResponse({'message': f'Maildate received: {maildate}'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def make_all_preport(request):
    queryset=PurchaseReport.objects.all().values()
    df = pd.DataFrame(list(queryset))
    random_number = random.randint(100, 999)
    file_excel_path=f"media/PurchaseReport_{random_number}.xlsx"
    df.to_excel(file_excel_path, index=False)
    response = FileResponse(open(file_excel_path, 'rb'), as_attachment=True, filename=f"PurchaseReport_{random_number}.xlsx")
    return response


@csrf_exempt
def convert_to_report(request):
    maildate=cache.get('invoicelist_date')
    print("maildate:",maildate)
    target_folder =  os.path.join(settings.MEDIA_ROOT,'out_invoice',maildate)
    if not os.path.exists(target_folder):
        return {'status': 'error', 'message': 'Folder does not exist'}
    for i in cache.get('invoicelist'):
        filepath=i['download_excel']
        filepath=filepath.lstrip('/')
        sheets = pd.read_excel(filepath,dtype=str)
        if sheets['supplier'].iloc[0]=="wurth":
            PurchaseReportServices.add_DataFrame_to_WurthReport(sheets)
        
        elif sheets['supplier'].iloc[0]=="John_McGrath":
            PurchaseReportServices.add_DataFrame_to_McGrathReport(sheets)
        
        elif sheets['supplier'].iloc[0]=="YHI AUSTRALIA":
            PurchaseReportServices.add_DataFrame_to_YhiaustraliaReport(sheets)
           
    wurth_record=WurthReport.objects.filter(maildate=cache.get("invoicelist_date")).values()
    John_McGrath_report=McGrathReport.objects.filter(maildate=maildate).values()
    YHI_report=YhiaustraliaReport.objects.filter(maildate=cache.get("invoicelist_date")).values()
    if wurth_record.exists():
        supplier_name="wurth"
        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
        col_map_list=[]
        for i in col_map:
            col_map_list.append(i)
        status=generatereport(wurth_record,col_map_list,case_col,supplier_name)

    if John_McGrath_report.exists():
        supplier_name="John_McGrath"
        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
        col_map_list=[]
        for i in col_map:
            col_map_list.append(i)
        status=generatereport(John_McGrath_report,col_map_list,case_col,supplier_name)
        print("generated:",status)

    if YHI_report.exists():
        supplier_name="YHI AUSTRALIA"
        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
        col_map_list=[]
        for i in col_map:
            col_map_list.append(i)
        status=generatereport(YHI_report,col_map_list,case_col,supplier_name)
    queryset=PurchaseReport.objects.filter(date=cache.get("invoicelist_date")).values()
    df = pd.DataFrame(list(queryset))
    random_number = random.randint(100, 999)
    file_excel_path=f"media/PurchaseReport_{random_number}.xlsx"
    df.to_excel(file_excel_path, index=False)
    response = FileResponse(open(file_excel_path, 'rb'), as_attachment=True, filename=f"PurchaseReport_{random_number}.xlsx")
    return response



@csrf_exempt
def upload_excel(request):
    if request.method == 'POST':
        files = request.FILES.getlist('files')
        maildate=cache.get('invoicelist_date')
        if maildate:
            if files:
                target_folder =  os.path.join(settings.MEDIA_ROOT,'out_invoice',maildate)
                for f in files:
                    if f.name.endswith('.xlsx'):
                        path = os.path.join(target_folder, f.name)
                        if os.path.exists(path):
                            os.remove(path)
                        with open(path, 'wb+') as destination:
                            for chunk in f.chunks():
                                destination.write(chunk)
            else:
                return JsonResponse({"status": "success",'message': 'Files ready to convert'})

        return JsonResponse({"status": "success",'message': 'Files uploaded successfully'})
    

@csrf_exempt
def generate_p_reportview(request):
     if request.method == "POST":
        data = json.loads(request.body)
        records=data.get('records')
        inputDate=data.get('inputDate')
        df=pd.DataFrame(records)
        random_number = random.randint(100, 999)
        file_excel_path=f"media/PurchaseReport_{random_number}.xlsx"
        df.to_excel(file_excel_path, index=False)
        
        return JsonResponse({
            "status": "success",
            "download_url": f"{file_excel_path}"
        })
        # response = FileResponse(open(file_excel_path, 'rb'), as_attachment=True, filename="PurchaseReport_{random_number}.xlsx")
        # return response

        generate_report_to_xlsl(records,inputDate)
        return JsonResponse({"status": "success", "message": "Records generated."}, status=200)

@csrf_exempt
def show_mail_dates_report(request):
    if request.method == "POST":
        input_date = request.POST.get('inputdate')
        supplier_id = request.POST.get('supplier_select')
        if supplier_id:
            print("has supplier id")
            Suppliername = Supplier.objects.filter(id=supplier_id).first().supplier_name
            p_report=PurchaseReport.objects.filter(supplier=Suppliername,date=input_date).values()
            response_data = {
            'purchase_records': list(p_report),
            'supplier_table_name': Suppliername
                }
            return JsonResponse(response_data, safe=False)
        else:
            print("has not supplier id")
            p_report=PurchaseReport.objects.filter(date=input_date).values()
            response_data = {
            'purchase_records': list(p_report),
            'supplier_table_name': "no supplier"
                }
            return JsonResponse(response_data, safe=False)


@csrf_exempt
def generate_report_reportview(request):
    if request.method == "POST":
        data = json.loads(request.body)
        records=data.get('records')
        supplier_name=data.get('supplier_table_name')
        supplier=Supplier.objects.filter(supplier_name=supplier_name).first()
        col_map=ColumnMapping.objects.filter(supplier_col=supplier_name).values()
        case_col=list(CaseModel.objects.filter(supplier=supplier).values())
        col_map_list=[]
        for i in col_map:
            col_map_list.append(i)
        # print("column mapping",list(col_map))
        # print("case_col",list(case_col))
        # print("supplier_name",type(supplier))
        status=generatereport(records,col_map_list,case_col,supplier_name)
        if status ==True:
            print('status:',status)
            return JsonResponse({"status": "success", "message": "Records generated."}, status=200)
          
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405) 
    

@csrf_exempt
def delete_case(request):
    if request.method == 'POST':
        minvalue = request.POST.get('minvalue')
        maxvalue = request.POST.get('maxvalue')
        actual_price=request.POST.get('actual_price')
        profit=request.POST.get('profit')
        selling_price_exc_gst=request.POST.get('selling_price_exc_gst')
        gst=request.POST.get('gst')
        selling_price_inc_gst=request.POST.get('selling_price_inc_gst')
        print("minvalue:",minvalue)
        print("maxvalue:",maxvalue)
        print("actual_price",actual_price)
        print("profit",profit)
        print("selling_price_exc_gst",selling_price_exc_gst)
        print("gst",gst)
        print("selling_price_inc_gst",selling_price_inc_gst)
        case=CaseModel.objects.filter(minvalue=minvalue,maxvalue=maxvalue,actual_price=actual_price,profit=profit,
                                 selling_price_exc_gst=selling_price_exc_gst,gst=gst,selling_price_inc_gst=selling_price_inc_gst
                                ) 
        case.delete()
        return JsonResponse({'success': True})


@csrf_exempt
def add_case_data(request):
    if request.method == 'POST':
        try:
            supplier_id = request.POST.get('supplier_id')
            supplier = Supplier.objects.filter(id=supplier_id).first()
            minvalue = request.POST.get('minvalue')
            maxvalue = request.POST.get('maxvalue')
            actual_price=request.POST.get('actual_price')
            profit = request.POST.get('profit')
            selling_price_exc_gst = request.POST.get('selling_price_exc_gst')
            gst = request.POST.get('gst')
            selling_price_inc_gst = request.POST.get('selling_price_inc_gst')
            if supplier:
                print("supplier_id",supplier.supplier_name,type(minvalue))
                caseobj = CaseModel.objects.create(
                    supplier=supplier,
                    minvalue=int(minvalue),
                    maxvalue=int(maxvalue),
                    actual_price=actual_price,
                    profit=profit,
                    selling_price_exc_gst=selling_price_exc_gst,
                    gst=gst,
                    selling_price_inc_gst=selling_price_inc_gst
                )
                caseobj.save()
                return JsonResponse({'status': 'success', 'message': 'Case created successfully'})
            return JsonResponse({'status': 'error', 'message': 'Supplier not found'}, status=400)
        except Supplier.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Invalid supplier ID'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)



@csrf_exempt
def showcases(request):
    if request.method == "POST":
        supplier_id = request.POST.get('supplier_id')
        try:
            supplier = Supplier.objects.get(id=supplier_id)
            cases = CaseModel.objects.filter(supplier=supplier).values()
            return JsonResponse({'message': 'Data fetched successfully', 'casedata': list(cases)})
        except Supplier.DoesNotExist:
            return JsonResponse({'message': 'Supplier not found'}, status=404)
    return JsonResponse({'message': 'Invalid request'}, status=400)
    

@csrf_exempt
def create_column_mapping(request):
    if request.method == "POST":
        data = request.POST
        supplierid=data.get('supplierid')
        supplier_col = data.get('supplier_col')
        date_col = data.get('date_col')
        part_description_col = data.get('part_description_col')
        part_number_col = data.get('part_number_col')
        trade_price_col = data.get('trade_price_col')
        total_count_col = data.get('total_count_col')
        purchase_count_col = data.get('purchase_count_col')
        total_price_col = data.get('total_price_col')
        actual_price_col = data.get('actual_price_col')
        profit_col = data.get('profit_col')
        selling_price_exc_gst_col = data.get('selling_price_exc_gst_col')
        gst_col = data.get('gst_col')
        selling_price_inc_gst_col = data.get('selling_price_inc_gst_col')
        print('suppliername:',supplierid)
        print('supplier_col:',supplier_col)
        if supplierid:
            try:
                supplier = Supplier.objects.filter(id=supplierid).first()
                if not supplier:
                    raise ValueError("Supplier not found.")
                supplier_name=supplier.supplier_name
                newmapping=ColumnMapping.objects.create(supplier_col=supplier_name,date_col=date_col,part_description_col=part_description_col,
                                                        part_number_col=part_number_col,trade_price_col=trade_price_col,total_count_col=total_count_col,purchase_count_col=purchase_count_col,
                                                        total_price_col=total_price_col,actual_price_col=actual_price_col,profit_col=profit_col,selling_price_exc_gst_col=selling_price_exc_gst_col,
                                                        gst_col=gst_col,selling_price_inc_gst_col=selling_price_inc_gst_col)
                newmapping.save()
                return JsonResponse({'status': 'success', 'message': 'columns mapping successfull'})
            except Exception as e:
                print(f"Unexpected error: {str(e)}")
                return JsonResponse({'status': 'error', 'message': 'Invalid operation'}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid operation'}, status=400)


@csrf_exempt
def get_suppliers_headers_value(request):
    if request.method == "POST":
        supplier_id=request.POST.get("selected_value")
        suppliername=Supplier.objects.filter(id=supplier_id).first().supplier_name
        col_map=ColumnMapping.objects.filter(supplier_col=suppliername).values()
        for i in col_map:
            print(i['supplier_col'])
        return JsonResponse({
        'status': 'success',
        'message': 'Data fetched successfully',
        'col_map': list(col_map)
        })


@csrf_exempt
def get_suppliers_headers(request):
    if request.method == "POST":
        suppliername=request.POST.get("selected_value")
        suppliername=Supplier.objects.filter(id=suppliername).first().supplier_name
        if suppliername:
            col_map=ColumnMapping.objects.filter(supplier_col=suppliername).values().first()
            if col_map:
                col_map_list = list(col_map.values())
                print(col_map_list)
        supplier_table_mapping = cache.get('supplier_table_name')
        supplier_table_name= supplier_table_mapping[suppliername]
        supplier_table_name= supplier_table_name[2:]
    return JsonResponse(supplier_table_name, safe=False)
    #return JsonResponse({'status': 'success', 'message': 'Data received'})


#get supplier names
@csrf_exempt
def get_suppliers(request):
    suppliers = Supplier.objects.all().values('id', 'supplier_name')
    return JsonResponse(list(suppliers), safe=False)



@csrf_exempt
def get_suppliers_column_map(request):
    print("supplier called:")
    suppliers = Supplier.objects.all().values('id', 'supplier_name')
    return JsonResponse(list(suppliers), safe=False)

@csrf_exempt
def add_supplier(request):
    if request.method == "POST":
        supplier_name = request.POST.get("supplier_name")
        if not supplier_name:
            return JsonResponse({'error': 'Missing supplier_name'}, status=400)
        if Supplier.objects.filter(supplier_name=supplier_name).exists():
            return JsonResponse({'error': 'Supplier already exists'}, status=409)
        supplier = Supplier.objects.create(supplier_name=supplier_name)
        return JsonResponse({'message': 'Supplier created', 'id': supplier.id, 'name': supplier.supplier_name})
    return JsonResponse({'error': 'Invalid method'}, status=405)

def settingsview(request):
    report_num_columns = cache.get('report_num_columns')
    report_num_columns={
        'report_num_columns':report_num_columns
    }
    return render(request,'settingsview.html',report_num_columns)

def update_report(request):   
    if request.method == "POST":
        try:     
            data = json.loads(request.body.decode('utf-8'))
            report_id = data.get("id")
            report = PurchaseReport.objects.get(id=report_id)
            report.date=data.get('date',report.date)
            report.supplier=data.get('supplier',report.supplier)
            report.part_description=data.get('part_description',report.part_description)
            report.part_number=data.get('part_number',report.part_number)
            report.trade_price=data.get('trade_price',report.trade_price)
            report.total_count=data.get('total_count',report.total_count)
            report.purchase_count=data.get('purchase_count',report.purchase_count)
            report.total_price=data.get('total_price',report.total_price)
            report.actual_price=data.get('actual_price',report.actual_price)
            report.profit=data.get('profit',report.profit)
            report.selling_price_exc_gst=data.get('selling_price_exc_gst',report.selling_price_exc_gst)
            report.gst=data.get('gst',report.gst)
            report.selling_price_inc_gst=data.get('selling_price_inc_gst',report.selling_price_inc_gst)
            report.save()
        except PurchaseReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)
        
        # print('report id:',report_id)   
        # print("report_date",report_date) 
        # print("report_supplier",report_supplier)
        # print("report_part_description",report_part_description)
        # print("report_part_number",report_part_number)
        # print("report_trade_price",report_trade_price)
        # print("report_total_count",report_total_count)
        # print("report_purchase_count",report_purchase_count)
        # print("report_total_price",report_total_price)
        # print("report_actual_price",report_actual_price)
        # print("report_profit",report_profit)
        # print("report_selling_price_exc_gst",report_selling_price_exc_gst)
        # print("report_gst",report_gst)
        # print("report_selling_price_inc_gst",report_selling_price_inc_gst)
        # report = PurchaseReport.objects.get(id=report_id)
        # try:
        #     report = PurchaseReport.objects.get()
        # except PurchaseReport.DoesNotExist:
        #     return JsonResponse({'error': 'Not found'}, status=404)
        return JsonResponse({'status': 'success'}, status=200)


def open_data_sheet(request,urldata):
    # try:
    #     data = list(InvoiceInfo.objects.filter(in_invoice_url=urldata).values())
    #     out_invoice_path = data[0]['ou_invoice_url']
    # except IndexError:
    #     out_invoice_path = None
    
    # df = pd.read_excel( out_invoice_path )
    # headerlist=df.columns.to_list()
    url_path = urldata.replace(settings.MEDIA_ROOT, settings.MEDIA_URL).replace("\\", "/")
    url_path = urldata.replace('media//','media/')
    return render(request, 'dataview.html', {'pdf_url': url_path})


@csrf_exempt
def show_mail_dates(request):
    if request.method == "POST":
        input_date = request.POST.get('inputdate')
        supplier_id = request.POST.get('supplier_select')
        Suppliername = Supplier.objects.filter(id=supplier_id).first()
        if not Suppliername:
            return JsonResponse({'status': 'error', 'message': 'Supplier not found'}, status=404)
        supp_dict = cache.get("suppliertable_modelname")
        if not supp_dict or Suppliername.supplier_name not in supp_dict:
            return JsonResponse({'status': 'error', 'message': 'Supplier model not found in cache'}, status=404)
        supplier_model_name = supp_dict[Suppliername.supplier_name]
        supplier_model = globals().get(supplier_model_name)
        if not supplier_model:
            return JsonResponse({'status': 'error', 'message': 'Model class not found'}, status=404)
        purchase_records_qs = supplier_model.objects.filter(maildate=input_date).values()
        purchase_records = list(purchase_records_qs)  # Convert queryset to list
        response_data = {
            'purchase_records': purchase_records,
            'supplier_table_name': Suppliername.supplier_name
        }
        return JsonResponse(response_data, safe=False)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


import importlib
@csrf_exempt
def save_record(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            update=data['supplier']
            if  update == "wurth":
                service_repository.updateWurthReport(data)
            elif update == "John_McGrath":
                service_repository.updateMcGrathReport(data)
            elif update == "YHI AUSTRALIA":
                service_repository.updateYhiaustraliaReport(data)            
            # update='update'+update
            # mod = importlib.import_module('service_repository')
            # func = getattr(mod,update)
            # print(update)
            # func(data)
            #service_repository.updateWurthReport(data)
            # record_id = data.get('id')
            # supplier_name = data.get('supplier_table_name')  # Must be passed
            # supp_dict = cache.get("suppliertable_modelname")
            # supplier_model = globals()[supp_dict[supplier_name]]
            # instance = supplier_model.objects.get(id=record_id)
            # for key, value in data.items():
            #     if hasattr(instance, key):
            #         setattr(instance, key, value)
            # instance.save()
            return JsonResponse({"status": "success", "message": "Record updated."})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)


@csrf_exempt
def delete_record(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            record_id = data.get('id')
            supplier_table_name = data.get('supplier_table_name')
            if  supplier_table_name == "wurth":
                instance=WurthReport.objects.get(id=record_id)
                instance.delete()
            elif supplier_table_name == "John_McGrath":
                instance=McGrathReport.objects.get(id=record_id)
                instance.delete()
            elif supplier_table_name == "YHI AUSTRALIA":
                instance=YhiaustraliaReport.objects.get(id=record_id)
                instance.delete()
            return JsonResponse({"status": "success", "message": "Record deleted."})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)


@csrf_exempt
def collect_mail_invoice(request):
    if request.method == 'POST':
            user_input = request.POST.get('maildata')
            if user_input:
                if checkmail_is_scrapped(user_input):
                    return JsonResponse({'status': 'error', 'message': 'mail is already scrapped!'}, status=400)
                #cache.set('invoicelist',[])
                target_date = datetime.strptime(user_input, "%Y-%m-%d").date()
                status,mail_subject,folder_date_path,attachments=MailAutomationClass.fetch_emails_on_date(target_date)
                if not os.path.isdir(folder_date_path):
                   return JsonResponse({'status': 'success', 'message':f"{user_input}Mail hasn't invoices" })
                pdf_files = [os.path.join(folder_date_path,f)  for f in os.listdir(folder_date_path) if os.path.isfile(os.path.join(folder_date_path, f)) and f.lower().endswith('.pdf')]
                dfsets=[]
                for i in pdf_files:
                    dfsets.append(UtilityClasses.scrap(i,user_input,mail_subject))
                    #UtilityClasses.scrap(i,user_input,mail_subject)
                if status:  
                    invoice_list=cache.get('invoicelist')
                    #cache.set('invoicelist_date',user_input)#setting cache maildate for listing upldading xlsl------------------------------------
                    ScrappedDate.objects.create(maildate=user_input)
                    return JsonResponse({'status': 'success', 'invoice_list': invoice_list},safe=False)
                else:
                    return JsonResponse({'status': 'error', 'message': 'No emails collected'}, status=400)
            else:
                return JsonResponse({'status': 'error', 'message': 'No date provided'}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def checkmail_is_scrapped(maildate):
    records = ScrappedDate.objects.filter(maildate=maildate)
    if records.exists():
        return True
    else:
        return False


# def upload_pdf(request):
#     if request.method == 'POST' and request.FILES.get('pdf_file'):
#         pdf = request.FILES['pdf_file']
#         fs = FileSystemStorage()
#         filename = fs.save(pdf.name, pdf)
#         cache.set('file_path',pdf.name)
#         value = cache.get('file_path')
#         return render(request, 'index.html', {'uploaded': True})
#     else:
#         return render(request, 'index.html', {'uploaded': False})

# InvoiceInfo.objects.all().delete()

#Supplier.objects.all().delete()

# i = Supplier.objects.filter(id=3).first()
# print(i.supplier_name)

#CalculationCase.objects.all().delete()

# field_names = [field.name for field in CalculationCase._meta.get_fields() if not field.many_to_many and not field.one_to_many]
# print(field_names)

# col=ColumnMapping.objects.all()
# for i in col:
#     print(i.actual_price_col)
# WurthReport.objects.all().delete()
# McGrathReport.objects.all().delete()
# YhiaustraliaReport.objects.all().delete()

# supp=Supplier.objects.all().values('supplier_name')
# print(supp)

#CaseModel.objects.all().delete()

# pp=PurchaseReport.objects.all()
# for i in pp:
#     print(i.id)
# pp=PurchaseReport.objects.all().values()
# print(pp)


PurchaseReport.objects.all().delete()
WurthReport.objects.all().delete()
McGrathReport.objects.all().delete()
YhiaustraliaReport.objects.all().delete()
InvoiceInfo.objects.all().delete()
ScrappedDate.objects.all().delete()
